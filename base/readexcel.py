import os
import sys

import openpyxl

def read_excel(excel_dir ,sheet_name,gn):
    '''读取excel'''

    filepath_final = get_project_path() + '\\' + excel_dir

    print(filepath_final)
    #加载目录
    ex=openpyxl.load_workbook(filepath_final)

    #获取sheet页
    sheet=ex[sheet_name]

    #打印表最大行和列
    # print(sheet.max_row,sheet.max_column)
    # print(sheet.cell(2,1).value)
    #循环行和列
    sheet_list=[]
    for row in range(2,sheet.max_row+1):
        row_list=[]
        for col in range(1,sheet.max_column+1):
            if sheet.cell(row,6).value == 'y':
                if  sheet.cell(row,2).value ==gn:
                    row_list.append(sheet.cell(row,col).value)
            else:
                pass
        if len(row_list) != 0:
            sheet_list.append(row_list)
    return sheet_list




def get_project_path(project_name='AutoTest'):
    """
    获取当前项目根路径
    :param project_name:
    :return: 根路径
    """
    PROJECT_NAME = 'selenium_project' if project_name is None else project_name
    project_path = os.path.abspath(os.path.dirname(__file__))
    root_path = project_path[:project_path.find("{}\\".format(PROJECT_NAME)) + len("{}\\".format(PROJECT_NAME))]
    # print('当前项目名称：{}\r\n当前项目根路径：{}'.format(PROJECT_NAME, root_path))
    return root_path

if __name__ == '__main__':
    print(read_excel('testcase.xlsx', '登录','登录'))
